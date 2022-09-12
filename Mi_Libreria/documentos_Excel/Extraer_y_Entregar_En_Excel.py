#debemos instalar para este archivo
#pip install openpyxl
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


#leyendo el archivo de excel
book= Workbook()
#ingresando el path de mi excel 
book=openpyxl.load_workbook('Mi_Libreria/documentos_Excel/Electrificadoras.xlsx')
#fijando la hoja activa
hoja=book.active

celdas = hoja['A1':'E8']
filas_Electrificadoras= hoja['A2':'A8']


##PROCESO DE INGRESO DE NUMEROS DE FACTURA DE LAS ESTACIONES A EL ARREGLO LISTADO DE FACTURAS

contador =0
listado_Id_Fact=[]
for i in filas_Electrificadoras:
    contador+=1
    if i[0].value=="CODENSA":
        listado_Id_Fact.append(celdas[contador][3].value)



##PROCESO DE SEPARACION DEL DV DEL ID DE LA FACTURA 
##INCORPORAR LOGICA DE BUSQUEDA MEDIANTE BOT
ticket_Respuestas_codensa=[]
contador =0 #Para llenar el ticket por el momento

num_Factura_completo=""
for celda in listado_Id_Fact:
        contador+=1

        num_Factura_completo=celda
        pos_guion = num_Factura_completo.find("-")
        numFactura=num_Factura_completo[0:pos_guion]
        digitoVer=num_Factura_completo[pos_guion+1:len(num_Factura_completo)]
        print(f" el numero de factura sin el DV es :{numFactura}")
        print(f"el digito de vrificacion es :{digitoVer}\n")
        
        ticket_Respuestas_codensa.append(f"contenido del ticket de respuesta codensa {contador}")


"""
##INGRESO DE LA RESPUESTA O TICKET A EL ARREGLO DE RESPUESTAS
##INCORPORAR LOGICA DE BUSQUEDA MEDIANTE BOT
contador =0
ticket_Respuestas_codensa=[]
for i in filas_Electrificadoras:
    contador+=1
    if i[0].value=="CODENSA":        
        ticket_Respuestas_codensa.append(f"ticket de respuesta codensa {contador}")

"""

        
##PROCESO DE LOCALIZACION DE COLUMNA PARA REGISTRAR LOS DATOS DE SALIDA ASOCIADOS AUNA RESPUESTA 
#para este caso se ingresa los datos almacenados en listado_Id_Fact=[]
contador =-1
b=0        
for i in filas_Electrificadoras:
    b+=1
    if i[0].value=="CODENSA":
        contador+=1        
        celdas[b][4].value=ticket_Respuestas_codensa[contador]
        print(celdas[b][4].value)

    

book.save('Mi_Libreria/documentos_Excel/Electrificadoras.xlsx')
"""

#seleccionando el rando de lectura de las celdas de excel de mi interes
celdas = hoja['D1':'D8']
print(celdas)

num_Factura_completo=""
for fila in celdas:
    for celda in fila:
        num_Factura_completo=celda.value
        print(num_Factura_completo)
        print(type(num_Factura_completo))
        pos_guion = num_Factura_completo.find("-")
        print(pos_guion)
        numFactura=num_Factura_completo[0:pos_guion]
        digitoVer=num_Factura_completo[pos_guion+1:len(num_Factura_completo)]
        print(numFactura)
        print(digitoVer)

"""